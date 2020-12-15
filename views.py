# Create your views here.
import csv
import os
import threading
import openpyxl
import pandas as pd
from bson import json_util
from rest_framework import status
from django.http import JsonResponse
import json
from LcmDataUpload.models import DataUploadSchema
from bson import ObjectId
from django.db import transaction

from LcmErrorHistoryUploadMapping.models import ErrorHistoryUploadMappingSchema
from LcmHistoryUploadMapping.models import HistoryUploadMappingSchema
from smbproject.Common.commonFunc import CommonFunc
from smbproject.Common import constants
import copy
import logging

from smbproject.Common.decorator import have_permission

common = CommonFunc()
logger = logging.getLogger(constants.NAME_LOG_APP)


@have_permission(common.permission(constants.TB_DATA_UPLOAD, constants.PERMISSION_READ))
def get_all(request):
    try:
        logger.info(constants.START_FUNCTION)
        db_data_upload = common.get_db(request, constants.TB_DATA_UPLOAD)
        # Validate request
        if not db_data_upload["status"]:
            return JsonResponse(data=common.to_dict(mess="User not valid!"),
                                status=status.HTTP_401_UNAUTHORIZED)

        lists_data = []
        limit = request.GET.get('limit', None)
        page = request.GET.get('page', None)
        if limit is not None and page is not None:
            limit = int(limit)
            page = int(page)
            lists_data = db_data_upload["data"].find({}).limit(limit).skip((page - 1) * limit)
        else:
            lists_data = db_data_upload["data"].find({})
        lists_data = json.loads(json_util.dumps(lists_data))

    except Exception as e:
        return JsonResponse(data=common.to_dict(mess=str(e)), status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    return JsonResponse(data=common.to_dict(data=lists_data), status=status.HTTP_200_OK)


@have_permission(common.permission(constants.TB_DATA_UPLOAD, constants.PERMISSION_READ))
def get_item(request):
    try:
        logger.info(constants.START_FUNCTION)
        db_data_upload = common.get_db(request, constants.TB_DATA_UPLOAD)
        # Validate request
        if not db_data_upload["status"]:
            return JsonResponse(data=common.to_dict(mess="User not valid!"),
                                status=status.HTTP_401_UNAUTHORIZED)
        body = json.loads(request.body)
        if "_id" not in body:
            return JsonResponse(data=common.to_dict(mess="Must have _id field!"),
                                status=status.HTTP_400_BAD_REQUEST)

        id_item = body['_id']
        item = db_data_upload["data"].find_one({"_id": id_item})
        item = json.loads(json_util.dumps(item))

    except Exception as e:
        logger.error(str(e))
        return JsonResponse(data=common.to_dict(mess=str(e)), status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    return JsonResponse(data=common.to_dict(data=item), status=status.HTTP_200_OK)


@have_permission(common.permission(constants.TB_DATA_UPLOAD, constants.PERMISSION_CREATE))
def upload_file(request):
    try:
        logger.info(constants.START_FUNCTION)
        db_data_upload = common.get_db(request, constants.TB_DATA_UPLOAD)
        # Validate request
        if not db_data_upload["status"]:
            return JsonResponse(data=common.to_dict(mess="User not valid!"),
                                status=status.HTTP_401_UNAUTHORIZED)

        if "file" not in request.FILES:
            return JsonResponse(data=common.to_dict(mess="Must have file field!"),
                                status=status.HTTP_400_BAD_REQUEST)
        if "campaign_id" not in request.POST:
            return JsonResponse(data=common.to_dict(mess="Must have campaign_id field!"),
                                status=status.HTTP_400_BAD_REQUEST)
        if "signal" not in request.POST:
            return JsonResponse(data=common.to_dict(mess="Must have signal field!"),
                                status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES["file"]
        campaign_id = request.POST['campaign_id']
        signal = request.POST['signal']
        id_item = str(ObjectId())

        file_name_full = str(file.name)
        file_name_store = os.path.join(constants.PATH_EXCEL_DATA_UPLOAD, id_item + "_" + file.name)
        # Validate extension file
        if not file_name_full.lower().endswith(('.csv', '.xlsx')):
            return JsonResponse(data=common.to_dict(mess="File extension must be .xlsx or .csv!"),
                                status=status.HTTP_400_BAD_REQUEST)

        # Read header file
        result_read = get_header_data_and_store_file(file, file_name_store, signal)
        item = {
            "_id": id_item,
            "campaign_id": campaign_id,
            "upload_data": [],
            "upload_result": constants.UPLOAD_RESULT_PROCESSING,
            "mapping_field": [],
            "mapping_mode": [],
            "file_name": id_item + "_" + file.name
        }
        DataUploadSchema().loads(json_util.dumps(common.convert_for_validate(item)))
        item = common.add_sub_infor(request, item)
        db_data_upload["data"].insert_one(item)

        result_upload = {
            "file_store": file_name_store,
            "file_name_original": file_name_full,
            "_id": id_item,
            "header_data": result_read["header_data"]
        }

        # Run async method
        thread = threading.Thread(target=update_data_upload,
                                  args=(request, id_item, result_read, constants.UPLOAD_RESULT_PROCESSING))
        thread.setDaemon(True)
        thread.start()

    except Exception as e:
        logger.error(str(e))
        return JsonResponse(data=common.to_dict(mess=str(e)), status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    return JsonResponse(data=common.to_dict(data=result_upload), status=status.HTTP_200_OK)


# Update upload data ( from file) or new upload data valid)
def update_data_upload(request, id_item, result_read, upload_result, new_upload_data=None, is_update_upload_data=True):
    logger.info(constants.START_FUNCTION)
    db_data_upload = common.get_db(request, constants.TB_DATA_UPLOAD)
    upload_data = []

    # upload data from file
    if is_update_upload_data:
        if new_upload_data is None:
            upload_data = get_upload_data_from_file(result_read)
        else:
            upload_data = new_upload_data

    item = {
        "_id": id_item,
        "campaign_id": "",
        "upload_data": upload_data,
        "upload_result": upload_result,
        "mapping_field": [],
        "mapping_mode": [],
        "file_name": ""
    }

    item = common.update_sub_info(request, item)
    DataUploadSchema().loads(json_util.dumps(common.convert_for_validate(item)))

    object_update = {
        "last_modify_time": item["last_modify_time"],
        "last_modify_by": item["last_modify_by"],
        "upload_result": item["upload_result"]
    }

    if is_update_upload_data:
        object_update.update({"upload_data": item["upload_data"]})

    db_data_upload["data"].update_one({"_id": item["_id"]},
                                      {"$set": object_update})


@have_permission(common.permission(constants.TB_DATA_UPLOAD, constants.PERMISSION_EDIT))
def upload_mapping(request):
    try:
        logger.info(constants.START_FUNCTION)
        db_data_upload = common.get_db(request, constants.TB_DATA_UPLOAD)
        # Validate request
        if not db_data_upload["status"]:
            return JsonResponse(data=common.to_dict(mess="User not valid!"),
                                status=status.HTTP_401_UNAUTHORIZED)
        body = json.loads(request.body)
        if "item" not in body:
            return JsonResponse(data=common.to_dict(mess="Must have item field!"),
                                status=status.HTTP_400_BAD_REQUEST)
        if "description_history_mapping" not in body:
            return JsonResponse(data=common.to_dict(mess="Must have description_history_mapping field!"),
                                status=status.HTTP_400_BAD_REQUEST)
        if "force_update_mapping" not in body:
            return JsonResponse(data=common.to_dict(mess="Must have force_update_mapping field!"),
                                status=status.HTTP_400_BAD_REQUEST)
        if "signal" not in body:
            return JsonResponse(data=common.to_dict(mess="Must have signal field!"),
                                status=status.HTTP_400_BAD_REQUEST)

        item = body["item"]
        if len(item["mapping_field"]) == 0 or len(item["mapping_mode"]) == 0:
            return JsonResponse(data=common.to_dict(mess="Must have choose at least 1 mapping field and 1 mapping mode!"),
                                status=status.HTTP_400_BAD_REQUEST)

        force_update_mapping = body["force_update_mapping"]
        description_history_mapping = body["description_history_mapping"]
        signal = body['signal']

        # Handle upload by auto upload
        if "mapping_auto_upload_id" in item:
            mapping_auto_upload_id = item["mapping_auto_upload_id"]
            db_mapping_auto_upload = common.get_db(request, constants.TB_MAPPING_AUTO_UPLOAD)
            item_mapping_auto_upload = db_mapping_auto_upload["data"].find_one({"_id": mapping_auto_upload_id})
            item["mapping_field"] = item_mapping_auto_upload["mapping_field"]
            item["mapping_mode"] = item_mapping_auto_upload["mapping_mode"]

        item_data_upload = db_data_upload["data"].find_one({"_id": item["_id"]})
        item["file_name"] = item_data_upload["file_name"]
        item["upload_data"] = []

        # Validate data mapping
        result_validate_mapping = validate_field_upload(request, item["mapping_field"], item["mapping_mode"])
        if not result_validate_mapping["status"]:
            return JsonResponse(data=common.to_dict(mess=result_validate_mapping["mess"]),
                                status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            # Normal update ( only update mapping and history)
            item = common.update_sub_info(request, item)
            DataUploadSchema().loads(json_util.dumps(common.convert_for_validate(item)))
            object_update = {
                "mapping_field": item["mapping_field"],
                "mapping_mode": item["mapping_mode"],
                "last_modify_time": item["last_modify_time"],
                "last_modify_by": item["last_modify_by"],
                "upload_result": constants.UPLOAD_RESULT_PROCESSING,
            }

            # handle main update
            db_data_upload["data"].update_one({"_id": item["_id"]}, {"$set": object_update})

            # case force update ( update uploads data valid)
            if force_update_mapping:
                result_read = None
                if not is_valid_result_validate_field(result_validate_mapping):
                    path_file = os.path.join(constants.PATH_EXCEL_DATA_UPLOAD, item_data_upload["file_name"])
                    result_read = get_result_read(path_file, signal)

                # Run async method handle validate data and update valid data
                thread = threading.Thread(target=validate_and_update_upload_data,
                                          args=(request, result_read, result_validate_mapping, force_update_mapping,
                                                item["_id"]))
                thread.setDaemon(True)
                thread.start()

            # handle update data history mapping
            if description_history_mapping is not None and description_history_mapping != "":
                db_history_upload_mapping = common.get_db(request, constants.TB_HISTORY_UPLOAD_MAPPING)

                item_history_upload_mapping = {
                    "_id": str(ObjectId()),
                    "description": description_history_mapping,
                    "campaign_id": item["campaign_id"],
                    "mapping_field": item["mapping_field"],
                    "mapping_mode": item["mapping_mode"]
                }
                item_history_upload_mapping = common.add_sub_infor(request, item_history_upload_mapping)
                HistoryUploadMappingSchema().loads(json_util.dumps(common.convert_for_validate(item_history_upload_mapping)))

                db_history_upload_mapping["data"].insert_one(item_history_upload_mapping)

    except Exception as e:
        logger.error(str(e))
        return JsonResponse(data=common.to_dict(mess=str(e)), status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    return JsonResponse(data=common.to_dict(data=item["_id"]), status=status.HTTP_200_OK)


# Get upload data from file excel or csv
def get_upload_data_from_file(result_read):
    logger.info(constants.START_FUNCTION)
    upload_data = []
    header_data = result_read["header_data"]

    # case file csv
    if result_read["type_extension"].lower().endswith('.csv'):
        df = result_read["object_read"]
        # row header
        for i, row in df.iterrows():
            # row body
            item_upload_data = {}
            for j, row_body in row.iteritems():
                items_cell = str(row_body).split(';')
                for index_cell, cell in enumerate(items_cell):
                    if index_cell < len(header_data):
                        if cell is None: cell = ""
                        item_cell = {header_data[index_cell]: cell}
                        item_upload_data.update(item_cell)
            upload_data.append(item_upload_data)

    # case file excel(xlsx)
    if result_read["type_extension"].lower().endswith('.xlsx'):
        worksheet = result_read["object_read"].active
        # iterating over the rows and
        for row in worksheet.iter_rows():
            item_upload_data = {}
            for cell in row:
                value_cell = str(cell.value)
                if cell.value is None: value_cell = ""
                # data body
                if not int(row[0].row) == 1:
                    item_cell = {header_data[int(cell.column) - 1]: value_cell}
                    item_upload_data.update(item_cell)
            if not int(row[0].row) == 1:
                upload_data.append(item_upload_data)
    return upload_data


# validate current upload data (file) and update valid upload data
def validate_and_update_upload_data(request, result_read, result_validate_mapping, force_update_mapping, id_item):
    # just update status
    if is_valid_result_validate_field(result_validate_mapping):
        update_data_upload(request, id_item, None, constants.UPLOAD_RESULT_SUCCESS, None, is_update_upload_data=False)
    # update upload data valid
    else:
        upload_data = get_upload_data_from_file(result_read)

        valid_data_upload = update_valid_upload_data(upload_data, result_validate_mapping["data"], force_update_mapping)
        upload_data = valid_data_upload["data"]["upload_data_result"]

        # update item Data upload with valid data
        update_data_upload(request, id_item, result_read, constants.UPLOAD_RESULT_SUCCESS, upload_data)

        #     insert all error record in excel to table
        db_error_history_upload_mapping = common.get_db(request, constants.TB_ERROR_HISTORY_UPLOAD_MAPPING)
        item_error_history_upload_mapping = {
            "_id": str(ObjectId()),
            "data_upload_id": id_item,
            "data_upload_error": valid_data_upload["data"]["upload_data_error"]
        }
        item_error_history_upload_mapping = common.add_sub_infor(request, item_error_history_upload_mapping)
        ErrorHistoryUploadMappingSchema().loads(json_util.dumps(common.convert_for_validate(item_error_history_upload_mapping)))

        db_error_history_upload_mapping["data"].insert_one(item_error_history_upload_mapping)


# Validate field (if mapping field, mapping mode miss required return fail else return object header validate
def validate_field_upload(request, mapping_field, mapping_mode):
    logger.info(constants.START_FUNCTION)

    # contain list type validate
    header_validate = {
        "header_required": []
    }
    is_valid_header = True
    fields_required = []

    db_business_info_config = common.get_db(request, constants.TB_BUSINESS_INFO_CONFIG)
    list_field = common.get_list_item_in_table(db_business_info_config["data"], "list_field")
    for field in list_field:
        add_field_one_time = False
        for item_mapping_field in mapping_field:
            if field["_id"] == item_mapping_field["field_id"]:
                if field["isrequire"]:
                    if not add_field_one_time:
                        header_validate["header_required"].append(item_mapping_field["upload_heading"])
                        add_field_one_time = True
                    if item_mapping_field["upload_heading"] is None or item_mapping_field["upload_heading"] == "":
                        is_valid_header = False
                        if not add_field_one_time: fields_required.append(field["field_label"])
        for item_mapping_mode in mapping_mode:
            if field["_id"] == item_mapping_mode["field_id"]:
                if field["isrequire"]:
                    if not add_field_one_time:
                        header_validate["header_required"].append(item_mapping_field["upload_heading"])
                        add_field_one_time = True
                    if item_mapping_mode["mode_id"] is None or item_mapping_mode["mode_id"] == "":
                        is_valid_header = False
                        if not add_field_one_time: fields_required.append(field["field_label"])

    if not is_valid_header:
        field_mess = ','.join(fields_required)
        return common.to_dict(status=False, mess="Field " + field_mess + " was required!")
    return common.to_dict(status=True, data=header_validate)


# Check validate of result validate field
def is_valid_result_validate_field(result_validate_field):
    header_validate = result_validate_field["data"]

    # validate required
    header_required = header_validate["header_required"]
    if len(header_required) > 0:
        return False

    return True


# Get header data excel and save file
def get_header_data_and_store_file(file, file_name_store, signal):
    logger.info(constants.START_FUNCTION)
    data_result = {
        "type_extension": "",
        "header_data": [],
        "object_read": {}
    }
    header_data = []
    if not os.path.exists(constants.PATH_EXCEL_DATA_UPLOAD):
        os.makedirs(constants.PATH_EXCEL_DATA_UPLOAD)

    if file.name.lower().endswith('.csv'):
        df = pd.read_csv(file)
        df.to_csv(file_name_store, index=False)
        for i, row in df.iterrows():
            header_data = str(row.axes[0].values[0]).split(signal)
            data_result["type_extension"] = '.csv'
            data_result["header_data"] = header_data
            data_result["object_read"] = df
            return data_result

    if file.name.lower().endswith('.xlsx'):
        wb = openpyxl.load_workbook(file)
        wb.save(file_name_store)
        worksheet = wb.active

        # iterating over the rows and
        for row in worksheet.iter_rows():
            if not int(row[0].row) == 1:
                data_result["type_extension"] = '.xlsx'
                data_result["header_data"] = header_data
                data_result["object_read"] = wb
                return data_result
            for cell in row:
                header_data.append(str(cell.value))


# Get result read from path file
def get_result_read(path_file, signal):
    logger.info(constants.START_FUNCTION)
    file = open(path_file, 'rb')

    data_result = {
        "type_extension": "",
        "header_data": [],
        "object_read": {}
    }
    header_data = []
    if not os.path.exists(constants.PATH_EXCEL_DATA_UPLOAD):
        os.makedirs(constants.PATH_EXCEL_DATA_UPLOAD)

    if file.name.lower().endswith('.csv'):
        df = pd.read_csv(file)
        for i, row in df.iterrows():
            header_data = str(row.axes[0].values[0]).split(signal)
            data_result["type_extension"] = '.csv'
            data_result["header_data"] = header_data
            data_result["object_read"] = df
            return data_result

    if file.name.lower().endswith('.xlsx'):
        wb = openpyxl.load_workbook(file)
        worksheet = wb.active

        # iterating over the rows and
        for row in worksheet.iter_rows():
            if not int(row[0].row) == 1:
                data_result["type_extension"] = '.xlsx'
                data_result["header_data"] = header_data
                data_result["object_read"] = wb
                return data_result
            for cell in row:
                header_data.append(str(cell.value))


# return result valid upload data in item data upload
def update_valid_upload_data(upload_data, result_validate_mapping, force_update_mapping):
    result_upload = {
        "upload_data_result": [],
        "upload_data_error": []
    }
    upload_data_result = copy.deepcopy(upload_data)
    upload_data_error = []
    for index, item in enumerate(upload_data):
        # delete item by rule header required
        for item_header in result_validate_mapping["header_required"]:
            if item_header not in item or item[item_header] is None or item[item_header] == "":
                if force_update_mapping:
                    upload_data_result[:] = [item_upload for item_upload in upload_data_result if item_upload != item]
                    upload_data_error.append(item)
                else:
                    return common.to_dict(status=False, mess="Field " + item_header + " in file is required!")
    result_upload["upload_data_result"] = upload_data_result
    result_upload["upload_data_error"] = upload_data_error
    return common.to_dict(status=True, data=result_upload)
